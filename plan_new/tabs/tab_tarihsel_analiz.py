import pandas as pd
import streamlit as st
from datetime import datetime
import matplotlib.pyplot as plt
import io
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import tempfile
import os

def tab_tarihsel_analiz(st, conn):
    st.subheader("📈 Tarihsel Uçuş Süre Analizi")

    df = pd.read_sql_query("SELECT plan_tarihi, sure, gorev_tipi FROM ucus_planlari", conn, parse_dates=["plan_tarihi"])
    if df.empty:
        st.warning("Veri bulunamadı.")
        return

    df["sure_saat"] = pd.to_timedelta(df["sure"]).dt.total_seconds() / 3600

    # === TARİH ARALIĞI SEÇİMİ ===
    min_date = df["plan_tarihi"].min()
    max_date = df["plan_tarihi"].max()
    col1, col2 = st.columns(2)
    with col1:
        tarih1 = st.date_input("Başlangıç Tarihi", min_value=min_date, max_value=max_date, value=min_date)
    with col2:
        tarih2 = st.date_input("Bitiş Tarihi", min_value=min_date, max_value=max_date, value=max_date)

    tarih1 = pd.to_datetime(tarih1)
    tarih2 = pd.to_datetime(tarih2)
    # Tarih aralığına göre filtrele
    mask = (df["plan_tarihi"] >= tarih1) & (df["plan_tarihi"] <= tarih2)
    df_aralik = df[mask].copy()

    # ---- ZEKİ OTOMATİK İSTATİSTİKLER ----
    toplam_saat = df_aralik["sure_saat"].sum()
    ort_gunluk = df_aralik.groupby("plan_tarihi")["sure_saat"].sum().mean()
    min_gun = df_aralik.groupby("plan_tarihi")["sure_saat"].sum().min()
    max_gun = df_aralik.groupby("plan_tarihi")["sure_saat"].sum().max()
    hic_ucus_olmayan_gunler = pd.date_range(tarih1, tarih2).difference(df_aralik["plan_tarihi"].unique())
    populer_tip = df_aralik.groupby("gorev_tipi")["sure_saat"].sum().idxmax() if not df_aralik.empty else None

    st.success(f"""
        **Tarih Aralığı: {tarih1.date()} – {tarih2.date()}**
        \n- Toplam Uçuş Süresi: **{toplam_saat:.2f} saat**
        \n- Günlük Ortalama: **{ort_gunluk:.2f} saat**
        \n- En yüksek uçulan gün: **{max_gun:.2f} saat**
        \n- En düşük (uçuş yapılan) gün: **{min_gun:.2f} saat**
        \n- En çok uçulan görev tipi: **{populer_tip}**
        \n- Hiç uçuş olmayan gün(ler): {'Yok' if len(hic_ucus_olmayan_gunler)==0 else ', '.join(str(g.date()) for g in hic_ucus_olmayan_gunler[:5])}
        """)

    # ==== Tüm Görev Tipleri: Günlük Toplam ====
    st.markdown("## 🔷 Toplam Uçuş Süresi (Günlük)")
    df_gunluk_total = df_aralik.groupby("plan_tarihi")["sure_saat"].sum().reset_index().sort_values("plan_tarihi")

    st.line_chart(
        df_gunluk_total.set_index("plan_tarihi"),
        height=300,
        use_container_width=True
    )
    st.dataframe(df_gunluk_total, use_container_width=True)

    # Görev tipleri
    gorev_tipleri = df_aralik["gorev_tipi"].dropna().unique().tolist()
    gorev_tipleri.sort()

    selected_tips = st.multiselect(
        "Gösterilecek Görev Tiplerini Seçin", 
        options=gorev_tipleri, 
        default=gorev_tipleri,
        key="tarihsel_gorev_tipleri"
    )

    # ==== Her Görev Tipi: Günlük Grafik ====
    for tip in selected_tips:
        st.markdown(f"---\n## {tip}")
        df_tip = df_aralik[df_aralik["gorev_tipi"] == tip]
        df_gunluk = df_tip.groupby("plan_tarihi")["sure_saat"].sum().reset_index().sort_values("plan_tarihi")

        st.line_chart(
            df_gunluk.set_index("plan_tarihi"),
            height=250,
            use_container_width=True
        )
        st.dataframe(df_gunluk, use_container_width=True)
        # Akıllı özet
        st.info(
            f"- Toplam: {df_tip['sure_saat'].sum():.2f} saat | "
            f"Ortalama: {df_gunluk['sure_saat'].mean():.2f} saat/gün | "
            f"Gün sayısı: {df_gunluk['plan_tarihi'].nunique()}"
        )

    # ==== Excel Rapor Butonu (Tarih aralığına göre, grafikler + özet) ====
    st.markdown("### 📥 Excel Raporu (Grafik+Özet) İndir")
    excel_rapor_grafikli_indir(st, df_aralik, df_gunluk_total, selected_tips, tarih1, tarih2)

# -- GRAFİKLİ EXCEL EXPORT, AKILLI ÖZETLERLE --
def excel_rapor_grafikli_indir(st, df_aralik, df_gunluk_total, selected_tips, tarih1, tarih2):
    import matplotlib.pyplot as plt
    import tempfile
    import os
    import io
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage

    tmpdir = tempfile.mkdtemp()
    # GENEL TOPLAM GRAFİK
    path_total = os.path.join(tmpdir, "tarihsel_total.png")
    plt.figure(figsize=(8,3))
    plt.plot(df_gunluk_total["plan_tarihi"], df_gunluk_total["sure_saat"], marker="o")
    plt.title("Toplam Uçuş Süresi (Günlük)")
    plt.xlabel("Tarih")
    plt.ylabel("Saat")
    plt.tight_layout()
    plt.grid(True, linestyle="--", alpha=0.5)
    plt.savefig(path_total)
    plt.close()

    # Her görev tipi için grafikler
    tip_graf_paths = []
    for tip in selected_tips:
        df_tip = df_aralik[df_aralik["gorev_tipi"] == tip]
        df_gunluk = df_tip.groupby("plan_tarihi")["sure_saat"].sum().reset_index()
        if df_gunluk.empty:
            continue
        pth = os.path.join(tmpdir, f"tip_{tip}.png")
        plt.figure(figsize=(7,2.5))
        plt.plot(df_gunluk["plan_tarihi"], df_gunluk["sure_saat"], marker=".")
        plt.title(f"Görev Tipi: {tip}")
        plt.xlabel("Tarih")
        plt.ylabel("Saat")
        plt.grid(True, linestyle="--", alpha=0.5)
        plt.tight_layout()
        plt.savefig(pth)
        plt.close()
        tip_graf_paths.append((tip, pth))

    # EXCEL YAZ
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        # Özet sheet
        summary = {
            "Tarih Başlangıç": tarih1.date(),
            "Tarih Bitiş": tarih2.date(),
            "Toplam Uçuş Saati": df_aralik["sure_saat"].sum(),
            "Ortalama Günlük": df_aralik.groupby("plan_tarihi")["sure_saat"].sum().mean(),
            "En çok uçulan görev tipi": df_aralik.groupby("gorev_tipi")["sure_saat"].sum().idxmax() if not df_aralik.empty else "-",
        }
        pd.DataFrame([summary]).to_excel(writer, index=False, sheet_name="Özet")

        # Genel toplam/günlük sheet
        df_gunluk_total.to_excel(writer, index=False, sheet_name="Günlük Toplam")
        # Her görev tipi ayrı sheet
        for tip in selected_tips:
            df_tip = df_aralik[df_aralik["gorev_tipi"] == tip]
            df_gunluk = df_tip.groupby("plan_tarihi")["sure_saat"].sum().reset_index()
            df_gunluk.to_excel(writer, index=False, sheet_name=f"Tip_{tip[:25]}")

    # Grafik ekle
    excel_buffer.seek(0)
    wb = load_workbook(excel_buffer)
    ws = wb["Günlük Toplam"]
    img = XLImage(path_total)
    ws.add_image(img, "E2")

    for tip, pth in tip_graf_paths:
        ws_tip = wb[f"Tip_{tip[:25]}"]
        ws_tip.add_image(XLImage(pth), "E2")

    # Özet sheet'e de akıllı istatistikler eklenebilir, şimdilik tablo şeklinde bırakıldı
    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)

    st.download_button(
        label="📥 Seçilen Tarih Aralığına Göre Grafikli Excel Raporu İndir",
        data=out_buf,
        file_name=f"ucus_analiz_{tarih1.date()}_{tarih2.date()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
